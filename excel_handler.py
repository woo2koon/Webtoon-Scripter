import os
import sys
import platform
import shutil
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from PySide6.QtWidgets import QFileDialog, QApplication
from widgets.message_box import CustomMessageBox

import config
from utils import restore_template, open_path

def export_to_excel(main_window):
    """
    엑셀 추출 비즈니스 로직을 처리하는 함수
    """
    if not main_window.current_episode: return
    e_path, _, _ = main_window.get_paths()
    
    try:
        main_window.save_char_data()
        main_window.save_script_data()
        
        export_episode_to_excel_core(main_window, e_path, main_window.current_title, main_window.current_episode, toast_target=main_window)
    except Exception as e:
        CustomMessageBox.critical(main_window, "오류", f"저장 중 오류 발생: {e}")

def export_episode_to_excel_core(parent_widget, e_path, title, episode, toast_target=None):
    """
    기존 호출 방식 유지용 (파일 대화창 포함)
    """
    default_filename = f"{title}_{episode}_스크립트.xlsx"
    # 마지막 저장 경로를 고려한 기본 경로 설정
    default_path = os.path.join(config.get_initial_dir(), default_filename)
    
    # [맥 네이티브 창 복구]
    options = QFileDialog.Option(0) if platform.system() == "Darwin" else QFileDialog.DontConfirmOverwrite
    save_path, _ = QFileDialog.getSaveFileName(parent_widget, "엑셀 저장", default_path, "Excel Files (*.xlsx)", options=options)
    
    if save_path:
        config.update_last_save_dir(save_path)
        # 중복 확인 얼럿 실행 (시스템 기본 테마)
        if os.path.exists(save_path):
            reply = CustomMessageBox.question(
                parent_widget,
                "파일 중복 확인",
                f"'{os.path.basename(save_path)}' 파일이 이미 존재합니다.\n기존 파일을 대체할까요, 아니면 새 이름으로 저장할까요?",
                ["덮어쓰기", "새 이름으로 저장", "취소"]
            )
            
            if reply == "덮어쓰기":
                pass 
            elif reply == "새 이름으로 저장":
                base, ext = os.path.splitext(save_path)
                counter = 1
                while os.path.exists(f"{base}({counter}){ext}"):
                    counter += 1
                save_path = f"{base}({counter}){ext}"
            else:
                return False

        return save_episode_to_excel_final(parent_widget, e_path, title, episode, save_path, toast_target)

def save_episode_to_excel_final(parent_widget, e_path, title, episode, save_path, toast_target=None):
    """
    실제 템플릿을 사용하여 엑셀 파일을 생성하고 저장하는 최종 핵심 로직
    (경로가 이미 결정된 경우 사용)
    """
    try:
        c_csv = os.path.join(e_path, "character_info.csv")
        s_csv = os.path.join(e_path, "script_data.csv")
        
        if os.path.exists(c_csv):
            char_df = pd.read_csv(c_csv, keep_default_na=False)
        else:
            char_df = pd.DataFrame(columns=['Character', 'Age', 'Gender', 'Role'])
            
        if os.path.exists(s_csv):
            script_df = pd.read_csv(s_csv, keep_default_na=False)
        else:
            script_df = pd.DataFrame(columns=['Character', 'Line'])
        
        if not os.path.exists(config.TEMPLATE_PATH): restore_template()
        wb = load_workbook(config.TEMPLATE_PATH)
        
        def copy_style(src, tgt):
            if src.has_style:
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.alignment = copy(src.alignment)
        
        if "캐릭터 정보" in wb.sheetnames:
            ws = wb["캐릭터 정보"]
            src_cells = [ws.cell(2, c) for c in range(1, 5)] if ws.max_row >= 2 else None
            if ws.max_row > 1: ws.delete_rows(2, ws.max_row)
            for i, row in char_df.iterrows():
                r = i + 2
                ws.cell(r, 1, row.get('Character', ''))
                ws.cell(r, 2, row.get('Age', ''))
                ws.cell(r, 3, row.get('Gender', ''))
                ws.cell(r, 4, row.get('Role', ''))
                if src_cells:
                    for c in range(1, 5): copy_style(src_cells[c-1], ws.cell(r, c))
        
        if "스크립트" in wb.sheetnames:
            ws = wb["스크립트"]
            src_cells = [ws.cell(2, c) for c in range(1, 4)] if ws.max_row >= 2 else None
            if ws.max_row > 1: ws.delete_rows(2, ws.max_row)
            dv = DataValidation(type="list", formula1="OFFSET('캐릭터 정보'!$A$2,0,0,COUNTA('캐릭터 정보'!$A:$A)-1,1)", allow_blank=True)
            ws.add_data_validation(dv)
            for i, row in script_df.iterrows():
                r = i + 2
                c1 = ws.cell(r, 1, row.get('Character', ''))
                c2 = ws.cell(r, 2, row.get('Line', ''))
                c3 = ws.cell(r, 3, "")
                dv.add(c1)
                if src_cells:
                    copy_style(src_cells[0], c1)
                    copy_style(src_cells[1], c2)
                    copy_style(src_cells[2], c3)

        # [수정] toast_target이 없으면 parent_widget을 기본으로 사용
        if not toast_target:
            toast_target = parent_widget
        
        # [추가] toast 속성이 직접 없으면 부모를 찾아 올라감 (ProjectManagementDialog 대응)
        actual_toast = None
        curr = toast_target
        while curr:
            if hasattr(curr, 'toast'):
                actual_toast = curr.toast
                break
            if hasattr(curr, 'parent'):
                curr = curr.parent()
            else:
                break

        if actual_toast:
            actual_toast.show_message("💾 엑셀 파일 저장 및 최적화 중...", 10000)
            if hasattr(actual_toast, 'opacity_effect'):
                actual_toast.opacity_effect.setOpacity(1.0)
            actual_toast.raise_()
            actual_toast.repaint()
            QApplication.processEvents()
            QApplication.processEvents()
        
        if sys.platform == "darwin":
            home = os.path.expanduser("~")
            temp_dir = os.path.join(home, "Library/Group Containers/UBF8T346G9.Office")
            if not os.path.exists(temp_dir):
                try: os.makedirs(temp_dir)
                except: temp_dir = tempfile.gettempdir()
        else:
            temp_dir = tempfile.gettempdir()

        temp_file_path = os.path.join(temp_dir, os.path.basename(save_path))
        wb.save(temp_file_path)
        wb.close()

        auto_save_success = False
        try:
            import xlwings as xw
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            if sys.platform == "win32":
                app.interactive = False
                app.screen_updating = False
            
            try:
                t_wb = app.books.open(temp_file_path)
                if sys.platform == "win32":
                    app.api.Visible = False 
                t_wb.save()
                t_wb.close()
                auto_save_success = True
            except:
                auto_save_success = False
            finally:
                app.quit()
                if sys.platform == "win32":
                    try: app.kill()
                    except: pass
        except ImportError:
            auto_save_success = False 

        final_nas_path = os.path.abspath(os.path.normpath(save_path))
        if os.path.exists(final_nas_path): os.remove(final_nas_path)
        shutil.move(temp_file_path, final_nas_path)
        
        if auto_save_success:
            if actual_toast:
                actual_toast.show_message("📄 엑셀파일 추출 완료", 3000)
        else:
            if actual_toast:
                actual_toast.show_message("⚠️ 추출 완료 (수동 저장 필요)", 4000)
            try:
                open_path(final_nas_path)
            except: pass
        
        return True
    except Exception as e:
        print(f"Excel Core Error: {e}")
        return False
